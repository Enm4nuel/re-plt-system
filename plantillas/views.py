from django.shortcuts import render, redirect
import os
import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import Font, Color

from .models import Template, TemplateLog

from db.db_manage import *

def loadData(building, coin, batch, rate, user):

	dm = DbManage()
	row = dm.connect1(building)

	leasid = []
	bldgid = []
	suitid = []
	occpname = []
	descripciones = []

	for i in range(len(row)):
		leasid.append(row[i][0])
		if "T-CAM" in row[i][1]:
			b = row[i][1].split(" ")
			bldgid.append(b[0])
		else:
			bldgid.append(row[i][1])
		suitid.append(row[i][2])
		occpname.append(row[i][3])

	t = Template.objects.all()
	print(t)
	for i in t:
		if i.bldgid == building and i.currcode == coin:
			descripciones.append(i.descrptn_name)

	generar_plantilla(user, batch, rate, coin, building, descripciones, leasid, bldgid, suitid, occpname)

	return 0

def generar_plantilla(username, batch, rate, coin, building, index = [], leasid = [], bldgid = [], suitid = [], occpname = []):
	d = {'LEASID': leasid, 'BLDGID': bldgid, 'SUITID': suitid, 'OCCPNAME': occpname}
	for i in range(len(index)):
		d[str(index[i])] = ""
	df = pd.DataFrame(data=d)
	df.to_excel("Plantilla_" + building + "_" + coin + "_" + batch + ".xlsx")

	TemplateLog.objects.create(username=username, bldgid=building, batch=batch, rate=rate)


def prueba_generar_plantilla(moneda, edificio, index = [], leasid = [], bldgid = [], suitid = [], occpname = []):

	wb = Workbook()
	ws = wb.active

	ws['A2'] = "Tasa"
	ws['B2'] = 57.99
	ws['A3'] = ""

	let = ['A', 'B', 'C', 'D', 'E', 'F', 'G']

	headers = ['LEASID', 'BLDGID', 'SUITID', 'OCCPNAME']

	for i in range(len(index)):
		headers.append(index[i])
	
	ws.append(headers)

	for i in range(len(leasid)):
		ws.append([leasid[i], bldgid[i], suitid[i], occpname[i]])
		

	# ESTILOS
	TITULO = Font(bold=True)

	row4 = ws.row_dimensions[4]
	row4.font = TITULO

	row1 = ws.row_dimensions[4]
	row1.font = Font()


	#ws['1'].font = Font(bold=True)

	wb.save("Plantilla" + edificio + "_" + moneda + ".xlsx")
